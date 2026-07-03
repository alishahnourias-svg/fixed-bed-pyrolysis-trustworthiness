# ============================================================
# Phase 12-C - Journal-Ready Figures and Controlled Results Skeleton
# Project:
# From Accuracy to Trustworthiness:
# Grouped Validation, Applicability-Domain Mapping,
# and Reliability-Aware Optimization for ML-Based
# Fixed-Bed Biomass Pyrolysis Yield Prediction
# ============================================================
#
# Purpose:
# This script refines Phase 12-B into a manuscript-facing package:
#   1) polished figure-ready data,
#   2) clearer PNG/SVG figures for journal drafting,
#   3) a controlled Results skeleton in Markdown,
#   4) a compact Excel workbook with final tables, captions, and guardrails.
#
# This script does NOT:
#   - train new models,
#   - run new optimization,
#   - validate true optima,
#   - turn bio-oil into an optimization target.
#
# Required input:
#   phase12b_manuscript_tables_and_figure_data.xlsx
# ============================================================

from pathlib import Path
import textwrap
import numpy as np
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import matplotlib.pyplot as plt


# ------------------------------------------------------------
# 0) Paths
# ------------------------------------------------------------

BASE_DIR = Path(".")
INPUT_XLSX = BASE_DIR / "phase12b_manuscript_tables_and_figure_data.xlsx"

OUTPUT_XLSX = BASE_DIR / "phase12c_journal_figures_and_results_package.xlsx"
OUTPUT_MD = BASE_DIR / "phase12c_results_skeleton.md"

FIG_DIR = BASE_DIR / "phase12c_journal_figures"
FIG_DIR.mkdir(exist_ok=True)

assert INPUT_XLSX.exists(), f"Missing input file: {INPUT_XLSX.resolve()}"

print("Loading:", INPUT_XLSX.resolve())


# ------------------------------------------------------------
# 1) Utilities
# ------------------------------------------------------------

VALIDATION_ORDER = [
    "Random K-fold",
    "Source-grouped",
    "Feedstock-grouped",
    "Family-grouped",
]

AD_FILTER_ORDER = [
    "all_predictions",
    "in_or_near_domain_only",
    "strict_in_domain_only",
]

AD_FILTER_LABELS = {
    "all_predictions": "All",
    "in_or_near_domain_only": "In/near-domain",
    "strict_in_domain_only": "Strict in-domain",
}

SHORT_GROUP_LABELS = {
    "agricultural_residue": "Agricultural residue",
    "shell_husk_kernel": "Shell/husk/kernel",
    "woody_biomass": "Woody biomass",
    "oilseed_residue": "Oilseed residue",
    "manure_sludge": "Manure/sludge",
    "Rice Husk": "Rice husk",
    "Rapeseed Stalk": "Rapeseed stalk",
    "Corncob": "Corncob",
    "Algae": "Algae",
    "Wood": "Wood",
    "Manure": "Manure",
    "Sunflower Bagasse": "Sunflower bagasse",
}


def read_sheet(name):
    return pd.read_excel(INPUT_XLSX, sheet_name=name)


def compact_float(df, digits=3):
    out = df.copy()
    for col in out.select_dtypes(include=[np.number]).columns:
        out[col] = out[col].round(digits)
    return out


def shorten_group(x):
    return SHORT_GROUP_LABELS.get(x, x)


def save_current_figure(stem):
    png = FIG_DIR / f"{stem}.png"
    svg = FIG_DIR / f"{stem}.svg"
    plt.tight_layout()
    plt.savefig(png, dpi=300, bbox_inches="tight")
    plt.savefig(svg, bbox_inches="tight")
    plt.close()
    return str(png), str(svg)


def add_bar_labels(ax, fmt="{:.2f}", orientation="vertical", padding=0.01):
    """
    Adds compact numeric labels. Works for positive bars.
    """
    if orientation == "vertical":
        ylim = ax.get_ylim()
        yspan = ylim[1] - ylim[0]
        for patch in ax.patches:
            height = patch.get_height()
            if np.isfinite(height):
                ax.text(
                    patch.get_x() + patch.get_width() / 2,
                    height + padding * yspan,
                    fmt.format(height),
                    ha="center",
                    va="bottom",
                    fontsize=8,
                    rotation=0,
                )
    else:
        xlim = ax.get_xlim()
        xspan = xlim[1] - xlim[0]
        for patch in ax.patches:
            width = patch.get_width()
            if np.isfinite(width):
                ax.text(
                    width + padding * xspan,
                    patch.get_y() + patch.get_height() / 2,
                    fmt.format(width),
                    ha="left",
                    va="center",
                    fontsize=8,
                )


def format_axes(ax):
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(axis="y", alpha=0.25)
    return ax


def format_axes_horizontal(ax):
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(axis="x", alpha=0.25)
    return ax


# ------------------------------------------------------------
# 2) Load Phase 12-B sheets
# ------------------------------------------------------------

T1 = read_sheet("T1_dataset_audit")
T2 = read_sheet("T2_validation_best")
T3 = read_sheet("T3_official_model_perf")
T4 = read_sheet("T4_uncertainty_90")
T5 = read_sheet("T5_ad_filtering")
T6 = read_sheet("T6_reliability_ranking")
T7 = read_sheet("T7_biochar_windows")
S1 = read_sheet("S1_warning_groups")
S2 = read_sheet("S2_uncertainty_assoc")
S3 = read_sheet("S3_ad_assoc")
S4 = read_sheet("S4_claims")
results_skeleton_old = read_sheet("Results_skeleton")


# ------------------------------------------------------------
# 3) Conservative claim-strength harmonization
# ------------------------------------------------------------
# Bio-oil should not be described as weak-to-moderate in the final narrative.
# It is a limited/cautionary screening case.

T3_clean = T3.copy()

if "Target" in T3_clean.columns and "Claim strength" in T3_clean.columns:
    mask_bio_oil = T3_clean["Target"].astype(str).str.lower().eq("bio-oil")
    T3_clean.loc[mask_bio_oil, "Claim strength"] = "limited_screening / cautionary"
    T3_clean.loc[mask_bio_oil, "Narrative guardrail"] = (
        "Use as a cautionary target; do not present as a robust optimization model."
    )

if "Narrative guardrail" not in T3_clean.columns:
    T3_clean["Narrative guardrail"] = ""

claim_strength_notes = pd.DataFrame([
    {
        "item": "bio_oil_claim_strength",
        "decision": "Use limited_screening / cautionary in the manuscript.",
        "reason": (
            "Bio-oil has high random-CV performance but weak grouped-validation performance "
            "and weak uncertainty-error association under grouped regimes."
        ),
    },
    {
        "item": "biochar_claim_strength",
        "decision": "Use moderate screening language only.",
        "reason": (
            "Biochar retains moderate grouped-validation performance and benefits more clearly "
            "from uncertainty and applicability-domain filtering."
        ),
    },
])


# ------------------------------------------------------------
# 4) Figure data preparation
# ------------------------------------------------------------

# Figure 2A/2B: validation performance
fig2_data = T2.copy()
fig2_data["Validation regime"] = pd.Categorical(
    fig2_data["Validation regime"],
    categories=VALIDATION_ORDER,
    ordered=True,
)
fig2_data = fig2_data.sort_values(["Validation regime", "Target"])

fig2_r2 = fig2_data.pivot(
    index="Validation regime",
    columns="Target",
    values="Mean R²",
).reindex(VALIDATION_ORDER)

fig2_rmse = fig2_data.pivot(
    index="Validation regime",
    columns="Target",
    values="RMSE",
).reindex(VALIDATION_ORDER)


# Figure 4A/4B: 90% uncertainty coverage and width
fig4_data = T4.copy()
fig4_data["Validation regime"] = pd.Categorical(
    fig4_data["Validation regime"],
    categories=VALIDATION_ORDER,
    ordered=True,
)
fig4_data = fig4_data.sort_values(["Validation regime", "Target"])

fig4_coverage = fig4_data.pivot(
    index="Validation regime",
    columns="Target",
    values="Absolute conformal coverage",
).reindex(VALIDATION_ORDER)

fig4_width = fig4_data.pivot(
    index="Validation regime",
    columns="Target",
    values="Absolute conformal mean width",
).reindex(VALIDATION_ORDER)


# Figure 5: Biochar AD filtering
fig5_data = T5[T5["Target"] == "Biochar"].copy()
fig5_data["AD filter label"] = fig5_data["AD filter"].map(AD_FILTER_LABELS).fillna(fig5_data["AD filter"])
fig5_data["Validation regime"] = pd.Categorical(
    fig5_data["Validation regime"],
    categories=VALIDATION_ORDER,
    ordered=True,
)
fig5_data["AD filter label"] = pd.Categorical(
    fig5_data["AD filter label"],
    categories=["All", "In/near-domain", "Strict in-domain"],
    ordered=True,
)
fig5_mae = fig5_data.pivot(
    index="Validation regime",
    columns="AD filter label",
    values="MAE",
).reindex(VALIDATION_ORDER)


# Figure 6: Yield-only OOD risk
fig6_data = T6.copy()
fig6_data["Group label"] = fig6_data["Group"].map(shorten_group)
fig6_data = fig6_data.sort_values("Yield-only out-of-domain fraction", ascending=True)

fig6_ood = fig6_data[[
    "Group label",
    "Yield-only out-of-domain fraction",
    "Reliability-aware out-of-domain fraction",
    "Reporting class",
]].copy()


# Figure 7: Biochar screening windows
fig7_data = T7.copy()
fig7_data["Group label"] = fig7_data["Group"].map(shorten_group)
fig7_data = fig7_data.sort_values("Mean 90% lower bound", ascending=True)
fig7_windows = fig7_data[[
    "Window level",
    "Group label",
    "Mean predicted yield",
    "Mean 90% lower bound",
    "Candidate in/near-domain fraction",
    "Selected in-domain fraction",
    "Selected near-domain fraction",
]].copy()


# ------------------------------------------------------------
# 5) Journal-draft figure generation
# ------------------------------------------------------------
# These are journal-draft figures, not final publisher-formatted artwork.
# They avoid long x-axis labels where possible.

figure_records = []

# Figure 2A: Mean R2
plt.figure(figsize=(8.0, 5.2))
ax = fig2_r2.plot(kind="bar", ax=plt.gca(), width=0.78)
format_axes(ax)
ax.set_title("Model performance under random and grouped validation")
ax.set_ylabel("Mean R²")
ax.set_xlabel("")
ax.set_ylim(0, max(0.95, np.nanmax(fig2_r2.values) * 1.15))
ax.tick_params(axis="x", rotation=25)
add_bar_labels(ax, fmt="{:.2f}", orientation="vertical")
png, svg = save_current_figure("fig2a_validation_mean_r2")
figure_records.append({
    "figure": "Figure 2A",
    "file_png": png,
    "file_svg": svg,
    "caption": (
        "Best-regime model performance under random and grouped validation. "
        "The drop from random to grouped validation indicates optimism in random K-fold evaluation, "
        "especially for bio-oil."
    ),
    "guardrail": "This figure shows best-by-regime performance, not necessarily the official model used downstream.",
})


# Figure 2B: RMSE
plt.figure(figsize=(8.0, 5.2))
ax = fig2_rmse.plot(kind="bar", ax=plt.gca(), width=0.78)
format_axes(ax)
ax.set_title("Prediction error under random and grouped validation")
ax.set_ylabel("RMSE")
ax.set_xlabel("")
ax.tick_params(axis="x", rotation=25)
add_bar_labels(ax, fmt="{:.1f}", orientation="vertical")
png, svg = save_current_figure("fig2b_validation_rmse")
figure_records.append({
    "figure": "Figure 2B",
    "file_png": png,
    "file_svg": svg,
    "caption": (
        "RMSE comparison across validation regimes. Grouped validation increases error, "
        "with a much stronger degradation for bio-oil than for biochar."
    ),
    "guardrail": "Do not interpret low random-CV RMSE as deployment-level reliability.",
})


# Figure 4A: Coverage
plt.figure(figsize=(8.0, 5.2))
ax = fig4_coverage.plot(kind="bar", ax=plt.gca(), width=0.78)
format_axes(ax)
ax.axhline(0.90, linestyle="--", linewidth=1)
ax.set_title("Coverage of 90% conformal intervals")
ax.set_ylabel("Empirical coverage")
ax.set_xlabel("")
ax.set_ylim(0, 1.08)
ax.tick_params(axis="x", rotation=25)
add_bar_labels(ax, fmt="{:.2f}", orientation="vertical")
png, svg = save_current_figure("fig4a_uncertainty_coverage_90")
figure_records.append({
    "figure": "Figure 4A",
    "file_png": png,
    "file_svg": svg,
    "caption": (
        "Empirical coverage of 90% absolute conformal intervals. Biochar remains close to nominal coverage, "
        "whereas bio-oil under-covers in source- and feedstock-grouped evaluation."
    ),
    "guardrail": "Coverage should be discussed together with interval width, not alone.",
})


# Figure 4B: Width
plt.figure(figsize=(8.0, 5.2))
ax = fig4_width.plot(kind="bar", ax=plt.gca(), width=0.78)
format_axes(ax)
ax.set_title("Width of 90% conformal intervals")
ax.set_ylabel("Mean interval width")
ax.set_xlabel("")
ax.tick_params(axis="x", rotation=25)
add_bar_labels(ax, fmt="{:.1f}", orientation="vertical")
png, svg = save_current_figure("fig4b_uncertainty_width_90")
figure_records.append({
    "figure": "Figure 4B",
    "file_png": png,
    "file_svg": svg,
    "caption": (
        "Mean width of 90% absolute conformal intervals. Grouped regimes require substantially wider intervals, "
        "highlighting deployment-like uncertainty inflation."
    ),
    "guardrail": "Wide intervals limit direct operational interpretation.",
})


# Figure 5: Biochar AD filtering MAE
plt.figure(figsize=(8.5, 5.2))
ax = fig5_mae.plot(kind="bar", ax=plt.gca(), width=0.78)
format_axes(ax)
ax.set_title("Biochar error after applicability-domain filtering")
ax.set_ylabel("MAE")
ax.set_xlabel("")
ax.tick_params(axis="x", rotation=25)
add_bar_labels(ax, fmt="{:.1f}", orientation="vertical")
png, svg = save_current_figure("fig5_biochar_ad_filtering_mae")
figure_records.append({
    "figure": "Figure 5",
    "file_png": png,
    "file_svg": svg,
    "caption": (
        "Effect of applicability-domain filtering on biochar MAE. In/near-domain and strict in-domain filters "
        "reduce error, but also retain fewer predictions."
    ),
    "guardrail": "Do not claim that AD filtering removes all error; it only reduces risk.",
})


# Figure 6: Yield-only OOD risk as horizontal chart
plt.figure(figsize=(8.8, 6.0))
plot6 = fig6_ood.set_index("Group label")[
    ["Yield-only out-of-domain fraction", "Reliability-aware out-of-domain fraction"]
]
ax = plot6.plot(kind="barh", ax=plt.gca(), width=0.78)
format_axes_horizontal(ax)
ax.set_title("Out-of-domain risk in yield-only versus reliability-aware ranking")
ax.set_xlabel("Out-of-domain fraction")
ax.set_ylabel("")
ax.set_xlim(0, 1.08)
add_bar_labels(ax, fmt="{:.2f}", orientation="horizontal")
png, svg = save_current_figure("fig6_yield_only_vs_reliability_aware_ood")
figure_records.append({
    "figure": "Figure 6",
    "file_png": png,
    "file_svg": svg,
    "caption": (
        "Out-of-domain fraction among top-ranked biochar scenarios. Yield-only ranking frequently selects "
        "out-of-domain candidates, whereas reliability-aware ranking excludes out-of-domain candidates by design."
    ),
    "guardrail": "Zero out-of-domain fraction does not mean experimentally validated optimum conditions.",
})


# Figure 7: Biochar screening windows, predicted vs lower bound
plt.figure(figsize=(8.8, 5.6))
plot7 = fig7_windows.set_index("Group label")[
    ["Mean predicted yield", "Mean 90% lower bound"]
]
ax = plot7.plot(kind="barh", ax=plt.gca(), width=0.75)
format_axes_horizontal(ax)
ax.set_title("Reliability-aware biochar screening windows")
ax.set_xlabel("Biochar yield (%)")
ax.set_ylabel("")
add_bar_labels(ax, fmt="{:.1f}", orientation="horizontal")
png, svg = save_current_figure("fig7_biochar_screening_windows")
figure_records.append({
    "figure": "Figure 7",
    "file_png": png,
    "file_svg": svg,
    "caption": (
        "Reliability-aware biochar screening windows summarized by mean predicted yield and conservative "
        "90% lower bound. These are screening targets, not experimentally validated optima."
    ),
    "guardrail": "Use as pre-experimental screening guidance only.",
})

figure_captions = pd.DataFrame(figure_records)


# ------------------------------------------------------------
# 6) Final Results skeleton text
# ------------------------------------------------------------

def md_escape(text):
    return str(text).replace("\n", " ").strip()


results_sections = [
    {
        "heading": "3.1 Dataset audit and target-specific availability",
        "body": (
            "After harmonization and target-specific filtering, the cleaned dataset contained 1,674 rows, "
            "with 1,548 usable records for biochar yield and 1,338 usable records for bio-oil yield. "
            "The dataset covered 68 harmonized feedstock labels and 124 primary source groups. "
            "Because target availability was not identical across outputs, biochar and bio-oil were evaluated separately."
        ),
        "table": "Table 1",
        "figure": "Figure 1",
        "guardrail": "Do not imply all rows contain both targets or complete features.",
    },
    {
        "heading": "3.2 Random cross-validation overestimates deployment-relevant performance",
        "body": (
            "Random K-fold validation produced high apparent performance for both targets, but this performance did not "
            "translate to grouped validation. For biochar, mean R² declined from approximately 0.823 under random K-fold "
            "to about 0.514–0.570 under grouped regimes. For bio-oil, the decline was much stronger, from approximately "
            "0.872 under random K-fold to about 0.065–0.128 under grouped regimes. These results indicate that random "
            "cross-validation substantially overestimates deployment-relevant generalization in this literature-derived dataset."
        ),
        "table": "Table 2",
        "figure": "Figure 2A–B",
        "guardrail": "State this as an optimism gap, not as proof that the models are useless.",
    },
    {
        "heading": "3.3 Target-dependent trustworthiness",
        "body": (
            "The grouped-validation results separate the two targets into different trustworthiness regimes. Biochar retained "
            "moderate screening value under grouped validation, whereas bio-oil behaved as a cautionary target: its high "
            "random-CV accuracy collapsed under source-, feedstock-, and family-grouped evaluation. For this reason, downstream "
            "uncertainty, applicability-domain, and scenario-ranking analyses were interpreted as reliability-aware screening for "
            "biochar and as diagnostic caution for bio-oil."
        ),
        "table": "Table 3",
        "figure": "Figure 2A–B",
        "guardrail": "Do not frame bio-oil as a reliable optimization target.",
    },
    {
        "heading": "3.4 Uncertainty calibration under grouped validation",
        "body": (
            "At the 90% nominal level, absolute conformal intervals for biochar maintained near-nominal empirical coverage "
            "across validation regimes, but grouped validation required substantially wider intervals. For bio-oil, intervals "
            "became much wider and under-covered in source- and feedstock-grouped evaluation. Thus, uncertainty calibration "
            "clarified deployment risk but did not eliminate the generalization problem, especially for bio-oil."
        ),
        "table": "Table 4",
        "figure": "Figure 4A–B",
        "guardrail": "Do not claim complete pointwise uncertainty quantification.",
    },
    {
        "heading": "3.5 Applicability-domain filtering",
        "body": (
            "Applicability-domain filtering improved biochar screening reliability by reducing error when predictions were "
            "restricted to in-domain or near-domain samples. The improvement was most useful for biochar; for bio-oil, the "
            "same filtering provided only limited protection, consistent with the weaker grouped-validation performance and "
            "weaker uncertainty-error association. Applicability-domain distance should therefore be interpreted as a risk "
            "filter rather than a full explanation of prediction error."
        ),
        "table": "Table 5",
        "figure": "Figure 5",
        "guardrail": "Do not claim AD filtering fully explains or fixes errors.",
    },
    {
        "heading": "3.6 Reliability-aware biochar scenario ranking",
        "body": (
            "Yield-only candidate ranking frequently selected scenarios with weak applicability-domain support. In several "
            "groups, including algae, corncob, wood, manure/sludge, agricultural residue, and woody biomass, the top yield-only "
            "candidates had high out-of-domain fractions. Reliability-aware ranking removed out-of-domain candidates and retained "
            "only in-domain or near-domain candidates, then summarized conservative biochar screening windows using the 90% lower "
            "bound. The final windows should be interpreted as pre-experimental screening targets, not as experimentally validated "
            "or industrial set-points."
        ),
        "table": "Table 6; Table 7",
        "figure": "Figure 6; Figure 7",
        "guardrail": "Do not call these true optimum conditions.",
    },
]

md_lines = []
md_lines.append("# Phase 12-C Results Skeleton")
md_lines.append("")
md_lines.append("> Scope: This is a controlled Results skeleton. It is not the final polished manuscript text.")
md_lines.append("> Guardrail: Use “screening windows” and “reliability-aware ranking”; avoid “true optimum conditions.”")
md_lines.append("")

for sec in results_sections:
    md_lines.append(f"## {sec['heading']}")
    md_lines.append("")
    md_lines.append(md_escape(sec["body"]))
    md_lines.append("")
    md_lines.append(f"**Use:** {sec['table']} | {sec['figure']}")
    md_lines.append("")
    md_lines.append(f"**Guardrail:** {sec['guardrail']}")
    md_lines.append("")

OUTPUT_MD.write_text("\n".join(md_lines), encoding="utf-8")

results_skeleton_final = pd.DataFrame(results_sections)


# ------------------------------------------------------------
# 7) Manuscript table captions and final table plan
# ------------------------------------------------------------

table_captions = pd.DataFrame([
    {
        "table": "Table 1",
        "title": "Dataset audit and target-specific data availability",
        "source_sheet": "T1_dataset_audit",
        "caption": (
            "Summary of the cleaned literature-derived fixed-bed pyrolysis dataset after harmonization. "
            "Usable record counts are target-specific."
        ),
        "guardrail": "Do not imply identical target availability for biochar and bio-oil.",
    },
    {
        "table": "Table 2",
        "title": "Best model performance under random and grouped validation regimes",
        "source_sheet": "T2_validation_best",
        "caption": (
            "Best-performing model/protocol combinations under random K-fold and grouped validation. "
            "The table quantifies the optimism gap between random and deployment-like grouped validation."
        ),
        "guardrail": "Best-by-regime performance is not the same as official downstream model selection.",
    },
    {
        "table": "Table 3",
        "title": "Official selected models and claim strength",
        "source_sheet": "T3_official_model_perf_clean",
        "caption": (
            "Official model choices used for downstream uncertainty, applicability-domain, and scenario-ranking analyses. "
            "The choices emphasize conservative interpretability and deployment-like defensibility."
        ),
        "guardrail": "Bio-oil should be described as limited/cautionary screening, not robust optimization.",
    },
    {
        "table": "Table 4",
        "title": "Ninety-percent conformal uncertainty calibration",
        "source_sheet": "T4_uncertainty_90",
        "caption": (
            "Empirical coverage and interval width of 90% conformal intervals under validation regimes."
        ),
        "guardrail": "Discuss coverage together with interval width.",
    },
    {
        "table": "Table 5",
        "title": "Effect of applicability-domain filtering on error and coverage",
        "source_sheet": "T5_ad_filtering",
        "caption": (
            "Effect of in-domain and near-domain filtering on error and interval coverage."
        ),
        "guardrail": "AD filtering reduces risk; it does not eliminate prediction error.",
    },
    {
        "table": "Table 6",
        "title": "Yield-only versus reliability-aware ranking for biochar candidate scenarios",
        "source_sheet": "T6_reliability_ranking",
        "caption": (
            "Comparison of out-of-domain risk in yield-only and reliability-aware ranking of candidate biochar scenarios."
        ),
        "guardrail": "Reliability-aware ranking excludes out-of-domain candidates by design; do not equate this with experimental validation.",
    },
    {
        "table": "Table 7",
        "title": "Reliability-aware biochar screening windows",
        "source_sheet": "T7_biochar_windows",
        "caption": (
            "Conservative biochar screening windows summarized from reliability-aware candidate rankings."
        ),
        "guardrail": "These windows are pre-experimental screening targets, not optimum industrial set-points.",
    },
])


# ------------------------------------------------------------
# 8) Quality gates
# ------------------------------------------------------------

quality_rows = []

for name, df in [
    ("T1_dataset", T1),
    ("T2_validation", T2),
    ("T3_official_clean", T3_clean),
    ("T4_uncertainty", T4),
    ("T5_ad_filtering", T5),
    ("T6_reliability", T6),
    ("T7_windows", T7),
    ("fig2_r2", fig2_r2),
    ("fig4_coverage", fig4_coverage),
    ("fig5_mae", fig5_mae),
    ("fig6_ood", fig6_ood),
    ("fig7_windows", fig7_windows),
]:
    quality_rows.append({
        "check": f"{name}_nonempty",
        "status": "pass" if len(df) > 0 else "fail",
        "detail": f"{len(df)} rows.",
    })

# Claim-strength check
bio_oil_claims = T3_clean[
    T3_clean.get("Target", pd.Series(dtype=str)).astype(str).str.lower().eq("bio-oil")
]

if len(bio_oil_claims) > 0 and "Claim strength" in bio_oil_claims.columns:
    too_positive = bio_oil_claims["Claim strength"].astype(str).str.contains("weak_to_moderate", case=False, na=False).any()
    quality_rows.append({
        "check": "bio_oil_claim_strength_harmonized",
        "status": "pass" if not too_positive else "warning",
        "detail": "Bio-oil claim strength harmonized to limited_screening / cautionary.",
    })

# Figure file checks
for rec in figure_records:
    quality_rows.append({
        "check": f"{rec['figure']}_png_exists",
        "status": "pass" if Path(rec["file_png"]).exists() else "fail",
        "detail": rec["file_png"],
    })
    quality_rows.append({
        "check": f"{rec['figure']}_svg_exists",
        "status": "pass" if Path(rec["file_svg"]).exists() else "fail",
        "detail": rec["file_svg"],
    })

quality_gates = pd.DataFrame(quality_rows)


# ------------------------------------------------------------
# 9) Save Phase 12-C workbook
# ------------------------------------------------------------

sheets = {
    "README": pd.DataFrame([
        {
            "field": "purpose",
            "value": "Journal-draft figures, final Results skeleton, captions, and guardrails for manuscript writing.",
        },
        {
            "field": "main_guardrail",
            "value": "Biochar windows are reliability-aware screening targets, not true optima.",
        },
        {
            "field": "bio_oil_guardrail",
            "value": "Bio-oil remains a cautionary target because grouped-validation performance is weak.",
        },
        {
            "field": "markdown_results_skeleton",
            "value": str(OUTPUT_MD),
        },
        {
            "field": "figure_directory",
            "value": str(FIG_DIR),
        },
    ]),
    "quality_gates": quality_gates,
    "T3_official_model_perf_clean": compact_float(T3_clean),
    "claim_strength_notes": claim_strength_notes,
    "figure_captions": figure_captions,
    "table_captions": table_captions,
    "results_skeleton_final": results_skeleton_final,
    "fig2_r2_data": compact_float(fig2_r2.reset_index()),
    "fig2_rmse_data": compact_float(fig2_rmse.reset_index()),
    "fig4_coverage_data": compact_float(fig4_coverage.reset_index()),
    "fig4_width_data": compact_float(fig4_width.reset_index()),
    "fig5_biochar_ad_mae_data": compact_float(fig5_mae.reset_index()),
    "fig6_ood_data": compact_float(fig6_ood),
    "fig7_biochar_windows_data": compact_float(fig7_windows),
    "T1_dataset_audit": compact_float(T1),
    "T2_validation_best": compact_float(T2),
    "T4_uncertainty_90": compact_float(T4),
    "T5_ad_filtering": compact_float(T5),
    "T6_reliability_ranking": compact_float(T6),
    "T7_biochar_windows": compact_float(T7),
    "S1_warning_groups": compact_float(S1),
    "S2_uncertainty_assoc": compact_float(S2),
    "S3_ad_assoc": compact_float(S3),
    "S4_claims": S4,
}

with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
    for sheet_name, table in sheets.items():
        table.to_excel(writer, sheet_name=sheet_name, index=False)


# ------------------------------------------------------------
# 10) Format workbook
# ------------------------------------------------------------

wb = load_workbook(OUTPUT_XLSX)

header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(color="FFFFFF", bold=True)
pass_fill = PatternFill("solid", fgColor="D9EAD3")
warning_fill = PatternFill("solid", fgColor="FFF2CC")
fail_fill = PatternFill("solid", fgColor="F4CCCC")
thin = Side(style="thin", color="D9E2F3")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for ws in wb.worksheets:
    ws.freeze_panes = "A2"

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

            if ws.title == "quality_gates" and cell.column == 2:
                val = str(cell.value).lower()
                if val == "pass":
                    cell.fill = pass_fill
                elif val == "warning":
                    cell.fill = warning_fill
                elif val == "fail":
                    cell.fill = fail_fill

            if ws.title == "S4_claims" and cell.column == 1:
                val = str(cell.value).lower()
                if val == "allowed":
                    cell.fill = pass_fill
                elif val == "avoid":
                    cell.fill = fail_fill

    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_length + 2, 10), 52)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.000"

wb.save(OUTPUT_XLSX)


# ------------------------------------------------------------
# 11) Console summary
# ------------------------------------------------------------

print("\nCreated:", OUTPUT_XLSX.resolve())
print("Created:", OUTPUT_MD.resolve())
print("Figures saved in:", FIG_DIR.resolve())

print("\nGenerated figures:")
for rec in figure_records:
    print(f"- {rec['figure']}: {rec['file_png']} | {rec['file_svg']}")

print("\nQuality gates:")
print(quality_gates.to_string(index=False))

print("\nNext files to review/upload:")
print("- phase12c_journal_figures_and_results_package.xlsx")
print("- phase12c_results_skeleton.md")
print("- selected PNG/SVG figures from phase12c_journal_figures/")
print("\nInterpretation reminder:")
print("These figures and skeleton support manuscript drafting; they do not validate true optimum conditions.")
