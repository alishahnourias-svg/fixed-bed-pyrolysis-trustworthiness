# ============================================================
# Phase 12-A - Manuscript-Ready Results Master Tables
# Project:
# From Accuracy to Trustworthiness:
# Grouped Validation, Applicability-Domain Mapping,
# and Reliability-Aware Optimization for ML-Based
# Fixed-Bed Biomass Pyrolysis Yield Prediction
# ============================================================
#
# Purpose:
# This script consolidates the final verified outputs from Phases 7-11
# into one manuscript-ready Excel workbook:
#
#   phase12_results_master_tables.xlsx
#
# It does NOT train new models.
# It does NOT run new optimization.
# It does NOT create new claims.
#
# It only packages the evidence into clean, traceable, paper-ready tables.
# ============================================================

from pathlib import Path
import numpy as np
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ------------------------------------------------------------
# 0) Paths and required inputs
# ------------------------------------------------------------

BASE_DIR = Path(".")

INPUTS = {
    "clean_data": "clean_audit_data_with_family.xlsx",
    "phase7_summary": "phase7_v2_summary_results.xlsx",
    "phase8_decisions": "phase8_official_model_decisions.xlsx",
    "phase9_calibration": "phase9_calibration_summary.xlsx",
    "phase9_assoc": "phase9_error_uncertainty_association.xlsx",
    "phase10_error_by_ad": "phase10_error_by_ad_class.xlsx",
    "phase10_filtering": "phase10_reliability_filtering_summary.xlsx",
    "phase10_assoc": "phase10_ad_error_association.xlsx",
    "phase11_family_windows": "phase11b3_v2_main_family_windows.xlsx",
    "phase11_case_windows": "phase11b3_v2_case_study_windows.xlsx",
    "phase11_risk": "phase11b3_v2_yield_only_risk_table.xlsx",
    "phase11_warning": "phase11b3_v2_excluded_or_warning_groups.xlsx",
}

OUTPUT_PATH = BASE_DIR / "phase12_results_master_tables.xlsx"


# ------------------------------------------------------------
# 1) Safe read utilities
# ------------------------------------------------------------

def read_excel_required(label, filename):
    path = BASE_DIR / filename
    if not path.exists():
        raise FileNotFoundError(
            f"Required input missing for {label}: {path.resolve()}"
        )
    return pd.read_excel(path)


def read_excel_optional(label, filename):
    path = BASE_DIR / filename
    if not path.exists():
        return None
    return pd.read_excel(path)


def compact_float(df, digits=3):
    out = df.copy()
    for col in out.select_dtypes(include=[np.number]).columns:
        out[col] = out[col].round(digits)
    return out


def safe_select(df, cols):
    existing = [c for c in cols if c in df.columns]
    return df[existing].copy()


def add_note_column(df, note):
    out = df.copy()
    out["interpretation_note"] = note
    return out


def duplicate_group_check(df, sheet_name):
    if df is None:
        return {
            "item": sheet_name,
            "status": "missing",
            "detail": "Table was not loaded.",
        }

    if not {"group_type", "group_value"}.issubset(set(df.columns)):
        return {
            "item": sheet_name,
            "status": "not_applicable",
            "detail": "No group_type/group_value columns.",
        }

    duplicate_count = int(df[["group_type", "group_value"]].duplicated().sum())

    if duplicate_count == 0:
        return {
            "item": sheet_name,
            "status": "pass",
            "detail": "No duplicated group_type/group_value rows.",
        }

    return {
        "item": sheet_name,
        "status": "fail",
        "detail": f"{duplicate_count} duplicated group rows detected.",
    }


# ------------------------------------------------------------
# 2) Load final project outputs
# ------------------------------------------------------------

clean_data = read_excel_optional("clean_data", INPUTS["clean_data"])

phase7 = read_excel_required("phase7_summary", INPUTS["phase7_summary"])
phase8 = read_excel_required("phase8_decisions", INPUTS["phase8_decisions"])
phase9_cal = read_excel_required("phase9_calibration", INPUTS["phase9_calibration"])
phase9_assoc = read_excel_required("phase9_assoc", INPUTS["phase9_assoc"])
phase10_ad = read_excel_required("phase10_error_by_ad", INPUTS["phase10_error_by_ad"])
phase10_filter = read_excel_required("phase10_filtering", INPUTS["phase10_filtering"])
phase10_assoc = read_excel_required("phase10_assoc", INPUTS["phase10_assoc"])
phase11_family = read_excel_required("phase11_family_windows", INPUTS["phase11_family_windows"])
phase11_cases = read_excel_required("phase11_case_windows", INPUTS["phase11_case_windows"])
phase11_risk = read_excel_required("phase11_risk", INPUTS["phase11_risk"])
phase11_warning = read_excel_required("phase11_warning", INPUTS["phase11_warning"])


# ------------------------------------------------------------
# 3) Dataset audit
# ------------------------------------------------------------

dataset_audit_rows = []

if clean_data is not None:
    dataset_audit_rows.append({
        "item": "cleaned_rows",
        "value": int(clean_data.shape[0]),
        "note": "Rows in clean_audit_data_with_family.xlsx.",
    })
    dataset_audit_rows.append({
        "item": "cleaned_columns",
        "value": int(clean_data.shape[1]),
        "note": "Columns in clean_audit_data_with_family.xlsx.",
    })

    if "biochar_yield" in clean_data.columns:
        dataset_audit_rows.append({
            "item": "usable_biochar_rows",
            "value": int(clean_data["biochar_yield"].notna().sum()),
            "note": "Rows with non-missing biochar yield.",
        })

    if "bio_oil_yield" in clean_data.columns:
        dataset_audit_rows.append({
            "item": "usable_bio_oil_rows",
            "value": int(clean_data["bio_oil_yield"].notna().sum()),
            "note": "Rows with non-missing bio-oil yield.",
        })

    if {"biochar_yield", "bio_oil_yield"}.issubset(clean_data.columns):
        dataset_audit_rows.append({
            "item": "paired_target_rows",
            "value": int(clean_data[["biochar_yield", "bio_oil_yield"]].notna().all(axis=1).sum()),
            "note": "Rows with both biochar and bio-oil yields.",
        })

    if "feedstock" in clean_data.columns:
        dataset_audit_rows.append({
            "item": "unique_feedstocks",
            "value": int(clean_data["feedstock"].dropna().nunique()),
            "note": "Unique harmonized feedstock labels.",
        })

    if "feedstock_family" in clean_data.columns:
        dataset_audit_rows.append({
            "item": "unique_feedstock_families",
            "value": int(clean_data["feedstock_family"].dropna().nunique()),
            "note": "Unique feedstock-family labels.",
        })

    if "source_group" in clean_data.columns:
        dataset_audit_rows.append({
            "item": "source_groups",
            "value": int(clean_data["source_group"].dropna().nunique()),
            "note": "Primary source groups used for source-grouped validation.",
        })

    if "source_pair_group" in clean_data.columns:
        dataset_audit_rows.append({
            "item": "source_pair_groups",
            "value": int(clean_data["source_pair_group"].dropna().nunique()),
            "note": "Reference-pair source groups.",
        })
else:
    dataset_audit_rows.append({
        "item": "clean_data_missing",
        "value": "not_available",
        "note": "clean_audit_data_with_family.xlsx was not found; dataset audit sheet is incomplete.",
    })

dataset_audit = pd.DataFrame(dataset_audit_rows)


# ------------------------------------------------------------
# 4) Validation core tables
# ------------------------------------------------------------

phase7_no_dummy = phase7.copy()

if "model" in phase7_no_dummy.columns:
    phase7_no_dummy = phase7_no_dummy[phase7_no_dummy["model"] != "Dummy_mean"].copy()

sort_cols = []
ascending = []

if "mean_r2" in phase7_no_dummy.columns:
    sort_cols.append("mean_r2")
    ascending.append(False)

if "mean_rmse" in phase7_no_dummy.columns:
    sort_cols.append("mean_rmse")
    ascending.append(True)

if not sort_cols:
    raise ValueError("phase7 summary must contain at least mean_r2 or mean_rmse.")

validation_best_by_regime = (
    phase7_no_dummy
    .sort_values(sort_cols, ascending=ascending)
    .groupby(["target", "validation_regime"], as_index=False)
    .head(1)
    .copy()
)

validation_best_by_regime = safe_select(
    validation_best_by_regime,
    [
        "target", "validation_regime", "protocol", "model",
        "n_folds", "total_test_rows",
        "mean_r2", "std_r2", "mean_rmse", "std_rmse", "mean_mae", "std_mae",
        "pooled_r2", "pooled_rmse", "pooled_mae",
        "mean_r2_drop_vs_random", "mean_rmse_increase_vs_random", "mean_mae_increase_vs_random",
        "pooled_r2_drop_vs_random", "pooled_rmse_increase_vs_random",
    ],
)

validation_best_by_regime = compact_float(validation_best_by_regime)

# Compact target-level contrast: random vs grouped best.
contrast_rows = []

for target, g in validation_best_by_regime.groupby("target"):
    random_rows = g[g["validation_regime"] == "random_kfold"]

    if len(random_rows) == 0:
        continue

    random_row = random_rows.iloc[0]

    for _, row in g.iterrows():
        regime = row["validation_regime"]

        contrast_rows.append({
            "target": target,
            "validation_regime": regime,
            "selected_protocol": row.get("protocol", np.nan),
            "selected_model": row.get("model", np.nan),
            "mean_r2": row.get("mean_r2", np.nan),
            "mean_rmse": row.get("mean_rmse", np.nan),
            "mean_mae": row.get("mean_mae", np.nan),
            "random_reference_mean_r2": random_row.get("mean_r2", np.nan),
            "random_reference_mean_rmse": random_row.get("mean_rmse", np.nan),
            "absolute_r2_drop_vs_random_best": (
                random_row.get("mean_r2", np.nan) - row.get("mean_r2", np.nan)
            ),
            "rmse_increase_vs_random_best": (
                row.get("mean_rmse", np.nan) - random_row.get("mean_rmse", np.nan)
            ),
            "interpretation": (
                "random_reference" if regime == "random_kfold"
                else "deployment_like_grouped_validation"
            ),
        })

validation_contrast = compact_float(pd.DataFrame(contrast_rows))


# ------------------------------------------------------------
# 5) Model decisions
# ------------------------------------------------------------

model_decisions = phase8.copy()


# ------------------------------------------------------------
# 6) Uncertainty core tables
# ------------------------------------------------------------

uncertainty_core = safe_select(
    phase9_cal,
    [
        "target", "role", "claim_strength", "protocol", "model", "validation_regime",
        "nominal_coverage", "n_predictions", "r2", "rmse", "mae",
        "coverage_abs_conformal", "mean_width_abs_conformal",
        "coverage_norm_conformal", "mean_width_norm_conformal",
        "mean_ensemble_std", "mean_abs_error",
    ],
)

uncertainty_core = compact_float(uncertainty_core)

uncertainty_assoc = safe_select(
    phase9_assoc,
    [
        "target", "role", "protocol", "model", "validation_regime", "n",
        "pearson_corr_ensemble_std_abs_error",
        "spearman_corr_ensemble_std_abs_error",
        "mean_abs_error", "mean_ensemble_std",
    ],
)

uncertainty_assoc = compact_float(uncertainty_assoc)


# ------------------------------------------------------------
# 7) Applicability-domain core tables
# ------------------------------------------------------------

ad_error_core = safe_select(
    phase10_ad,
    [
        "target", "role", "claim_strength", "protocol", "model",
        "validation_regime", "combined_ad_class", "n", "class_fraction",
        "mean_abs_error", "median_abs_error", "rmse",
        "mean_knn_distance", "mean_mahalanobis_distance",
    ],
)

ad_error_core = compact_float(ad_error_core)

ad_filtering_core = safe_select(
    phase10_filter,
    [
        "target", "role", "claim_strength", "protocol", "model",
        "validation_regime", "filter_name", "n_kept", "fraction_kept",
        "mean_abs_error", "median_abs_error", "rmse",
        "coverage_abs_90", "mean_width_abs_90",
        "coverage_norm_90", "mean_width_norm_90",
    ],
)

ad_filtering_core = compact_float(ad_filtering_core)

ad_assoc_core = safe_select(
    phase10_assoc,
    [
        "target", "role", "protocol", "model", "validation_regime", "n",
        "spearman_knn_abs_error", "pearson_knn_abs_error",
        "spearman_mahalanobis_abs_error", "pearson_mahalanobis_abs_error",
        "spearman_ensemble_std_abs_error",
        "mean_abs_error",
    ],
)

ad_assoc_core = compact_float(ad_assoc_core)


# ------------------------------------------------------------
# 8) Reliability-aware scenario tables
# ------------------------------------------------------------

reliability_ranking = compact_float(phase11_risk.copy())
biochar_family_windows = compact_float(phase11_family.copy())
biochar_case_windows = compact_float(phase11_cases.copy())
warning_groups = compact_float(phase11_warning.copy())


# ------------------------------------------------------------
# 9) Key manuscript numbers
# ------------------------------------------------------------

key_rows = []

def add_best_metric_row(target, regime, label):
    rows = validation_best_by_regime[
        (validation_best_by_regime["target"] == target)
        & (validation_best_by_regime["validation_regime"] == regime)
    ]
    if len(rows) == 0:
        return

    row = rows.iloc[0]

    key_rows.append({
        "section": "validation",
        "label": label,
        "target": target,
        "validation_regime": regime,
        "model": row.get("model", np.nan),
        "protocol": row.get("protocol", np.nan),
        "mean_r2": row.get("mean_r2", np.nan),
        "mean_rmse": row.get("mean_rmse", np.nan),
        "mean_mae": row.get("mean_mae", np.nan),
        "claim_use": "Core result for manuscript narrative.",
    })

for target in ["biochar", "bio_oil"]:
    add_best_metric_row(target, "random_kfold", f"{target}: best random-CV performance")
    add_best_metric_row(target, "source_group_kfold", f"{target}: best source-grouped performance")
    add_best_metric_row(target, "feedstock_group_kfold", f"{target}: best feedstock-grouped performance")
    add_best_metric_row(target, "family_group_kfold", f"{target}: best family-grouped performance")

# Phase 11 summary
key_rows.append({
    "section": "reliability_ranking",
    "label": "biochar family windows retained for main text",
    "target": "biochar",
    "validation_regime": "not_applicable",
    "model": "ExtraTrees",
    "protocol": "A_numeric_only",
    "mean_r2": np.nan,
    "mean_rmse": np.nan,
    "mean_mae": np.nan,
    "claim_use": f"{len(biochar_family_windows)} feedstock-family windows retained for main-text screening table.",
})

key_rows.append({
    "section": "reliability_ranking",
    "label": "biochar feedstock case-study windows retained",
    "target": "biochar",
    "validation_regime": "not_applicable",
    "model": "ExtraTrees",
    "protocol": "A_numeric_only",
    "mean_r2": np.nan,
    "mean_rmse": np.nan,
    "mean_mae": np.nan,
    "claim_use": f"{len(biochar_case_windows)} feedstock-level case-study windows retained.",
})

key_manuscript_numbers = compact_float(pd.DataFrame(key_rows))


# ------------------------------------------------------------
# 10) Table plan, figure plan, claims
# ------------------------------------------------------------

table_plan = pd.DataFrame([
    {
        "table_id": "Table 1",
        "proposed_title": "Dataset audit and target-specific data availability",
        "source_sheet": "dataset_audit",
        "main_or_supplement": "main",
        "purpose": "Establish data size, target availability, feedstock/source coverage.",
    },
    {
        "table_id": "Table 2",
        "proposed_title": "Best model performance under random and grouped validation regimes",
        "source_sheet": "validation_core",
        "main_or_supplement": "main",
        "purpose": "Show random-CV optimism and target-specific grouped generalization.",
    },
    {
        "table_id": "Table 3",
        "proposed_title": "Selected models and defensible claim strength",
        "source_sheet": "model_decisions",
        "main_or_supplement": "supplement_or_short_main",
        "purpose": "Document why biochar is retained for screening and bio-oil is limited.",
    },
    {
        "table_id": "Table 4",
        "proposed_title": "Conformal uncertainty calibration under validation regimes",
        "source_sheet": "uncertainty_core",
        "main_or_supplement": "main",
        "purpose": "Show coverage/width trade-offs and deployment-like uncertainty inflation.",
    },
    {
        "table_id": "Table 5",
        "proposed_title": "Applicability-domain filtering effects on error and coverage",
        "source_sheet": "ad_filtering",
        "main_or_supplement": "main",
        "purpose": "Show AD filtering is useful mainly for biochar.",
    },
    {
        "table_id": "Table 6",
        "proposed_title": "Yield-only versus reliability-aware candidate ranking for biochar screening",
        "source_sheet": "reliability_ranking",
        "main_or_supplement": "main",
        "purpose": "Show yield-only ranking often selects out-of-domain scenarios.",
    },
    {
        "table_id": "Table 7",
        "proposed_title": "Reliability-aware biochar screening windows",
        "source_sheet": "biochar_family + biochar_cases",
        "main_or_supplement": "main",
        "purpose": "Report conservative pre-experimental screening windows, not optima.",
    },
    {
        "table_id": "Supplementary tables",
        "proposed_title": "Full validation, AD class, warning-group, and excluded-group details",
        "source_sheet": "validation_all/ad_error_class/warning_groups",
        "main_or_supplement": "supplement",
        "purpose": "Keep main text focused while preserving transparency.",
    },
])

figure_plan = pd.DataFrame([
    {
        "figure_id": "Figure 1",
        "proposed_title": "Trustworthiness workflow",
        "source_sheet": "conceptual",
        "main_or_supplement": "main",
        "message": "From random CV to grouped validation, uncertainty, AD, and reliability-aware ranking.",
    },
    {
        "figure_id": "Figure 2",
        "proposed_title": "Random-CV optimism under grouped validation",
        "source_sheet": "validation_contrast",
        "main_or_supplement": "main",
        "message": "Performance drops from random to source/feedstock/family-grouped validation.",
    },
    {
        "figure_id": "Figure 3",
        "proposed_title": "Target-dependent trustworthiness: biochar versus bio-oil",
        "source_sheet": "validation_core + uncertainty_core",
        "main_or_supplement": "main",
        "message": "Biochar retains moderate screening value; bio-oil collapses under grouped validation.",
    },
    {
        "figure_id": "Figure 4",
        "proposed_title": "Uncertainty coverage and interval width across validation regimes",
        "source_sheet": "uncertainty_core",
        "main_or_supplement": "main",
        "message": "Group-aware evaluation increases interval width and reveals under-coverage risk.",
    },
    {
        "figure_id": "Figure 5",
        "proposed_title": "Error by applicability-domain class",
        "source_sheet": "ad_error_class",
        "main_or_supplement": "main",
        "message": "Out-of-domain biochar points are generally more error-prone; bio-oil AD signal is weaker.",
    },
    {
        "figure_id": "Figure 6",
        "proposed_title": "Yield-only versus reliability-aware biochar scenario ranking",
        "source_sheet": "reliability_ranking",
        "main_or_supplement": "main",
        "message": "Reliability-aware ranking removes out-of-domain candidates and reports conservative windows.",
    },
])

claims = pd.DataFrame([
    {
        "claim_type": "allowed",
        "claim": "Random cross-validation overestimates deployment-relevant performance in literature-derived pyrolysis ML datasets.",
        "supporting_sheets": "validation_core; validation_contrast",
    },
    {
        "claim_type": "allowed",
        "claim": "The optimism gap is much more severe for bio-oil than for biochar.",
        "supporting_sheets": "validation_core; uncertainty_core",
    },
    {
        "claim_type": "allowed",
        "claim": "Biochar retains moderate screening value under grouped validation, especially when paired with uncertainty and AD filters.",
        "supporting_sheets": "model_decisions; uncertainty_core; ad_filtering; biochar_family",
    },
    {
        "claim_type": "allowed",
        "claim": "Bio-oil should be treated as a cautionary case rather than a reliable optimization target in this dataset.",
        "supporting_sheets": "validation_core; uncertainty_core; ad_error_class",
    },
    {
        "claim_type": "allowed",
        "claim": "Reliability-aware ranking reduces out-of-domain yield-only candidates and produces conservative biochar screening windows.",
        "supporting_sheets": "reliability_ranking; biochar_family; biochar_cases",
    },
    {
        "claim_type": "avoid",
        "claim": "The true optimum fixed-bed pyrolysis conditions were determined.",
        "supporting_sheets": "none; this is not supported by the study design",
    },
    {
        "claim_type": "avoid",
        "claim": "The models are reliable for all feedstocks and all operating regimes.",
        "supporting_sheets": "contradicted by grouped validation and AD results",
    },
    {
        "claim_type": "avoid",
        "claim": "Bio-oil yield can be robustly optimized using the current model.",
        "supporting_sheets": "contradicted by bio-oil grouped-validation and uncertainty results",
    },
    {
        "claim_type": "avoid",
        "claim": "Applicability-domain distance fully explains prediction errors.",
        "supporting_sheets": "contradicted by weak/moderate AD-error correlations",
    },
    {
        "claim_type": "avoid",
        "claim": "The scenario windows are experimentally validated industrial set-points.",
        "supporting_sheets": "not supported; windows are pre-experimental screening targets only",
    },
])


# ------------------------------------------------------------
# 11) Quality gates
# ------------------------------------------------------------

quality_rows = []

for label, filename in INPUTS.items():
    path = BASE_DIR / filename
    if label == "clean_data":
        status = "pass" if path.exists() else "warning"
        detail = "Optional for dataset audit; missing file only makes dataset audit incomplete."
    else:
        status = "pass" if path.exists() else "fail"
        detail = "Required final output file."

    quality_rows.append({
        "check": f"input_exists::{label}",
        "status": status,
        "detail": f"{filename} | {detail}",
    })

for sheet_name, table in [
    ("phase11_family_windows", phase11_family),
    ("phase11_case_windows", phase11_cases),
    ("phase11_risk", phase11_risk),
    ("phase11_warning", phase11_warning),
]:
    result = duplicate_group_check(table, sheet_name)
    quality_rows.append({
        "check": f"duplicate_groups::{sheet_name}",
        "status": result["status"],
        "detail": result["detail"],
    })

quality_rows.append({
    "check": "phase11_main_family_count",
    "status": "pass" if len(phase11_family) >= 1 else "warning",
    "detail": f"{len(phase11_family)} main family rows retained.",
})

quality_rows.append({
    "check": "phase11_case_study_count",
    "status": "pass" if len(phase11_cases) >= 1 else "warning",
    "detail": f"{len(phase11_cases)} case-study rows retained.",
})

quality_rows.append({
    "check": "manuscript_language_guardrail",
    "status": "pass",
    "detail": "Use screening windows and reliability-aware ranking; avoid claiming true optima or industrial deployment.",
})

quality_gates = pd.DataFrame(quality_rows)


# ------------------------------------------------------------
# 12) Write workbook
# ------------------------------------------------------------

sheets = {
    "README": pd.DataFrame([
        {
            "field": "workbook_purpose",
            "value": "Manuscript-ready evidence package for Phase 12.",
        },
        {
            "field": "core_message",
            "value": (
                "Random CV is optimistic; grouped validation, uncertainty, AD mapping, "
                "and reliability-aware ranking provide a more defensible trustworthiness assessment."
            ),
        },
        {
            "field": "scope_warning",
            "value": (
                "Biochar windows are pre-experimental screening targets; they are not experimentally validated optima."
            ),
        },
        {
            "field": "bio_oil_warning",
            "value": (
                "Bio-oil results should be used as a cautionary trustworthiness case, not as an optimization recommendation."
            ),
        },
    ]),
    "quality_gates": quality_gates,
    "dataset_audit": dataset_audit,
    "validation_core": validation_best_by_regime,
    "validation_contrast": validation_contrast,
    "model_decisions": model_decisions,
    "uncertainty_core": uncertainty_core,
    "uncert_assoc": uncertainty_assoc,
    "ad_error_class": ad_error_core,
    "ad_filtering": ad_filtering_core,
    "ad_assoc": ad_assoc_core,
    "reliability_ranking": reliability_ranking,
    "biochar_family": biochar_family_windows,
    "biochar_cases": biochar_case_windows,
    "warning_groups": warning_groups,
    "key_numbers": key_manuscript_numbers,
    "table_plan": table_plan,
    "figure_plan": figure_plan,
    "claims": claims,
}

with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
    for sheet_name, table in sheets.items():
        table.to_excel(writer, sheet_name=sheet_name, index=False)


# ------------------------------------------------------------
# 13) Workbook formatting
# ------------------------------------------------------------

wb = load_workbook(OUTPUT_PATH)

header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(color="FFFFFF", bold=True)
subtle_fill = PatternFill("solid", fgColor="D9EAF7")
warning_fill = PatternFill("solid", fgColor="FFF2CC")
fail_fill = PatternFill("solid", fgColor="F4CCCC")
pass_fill = PatternFill("solid", fgColor="D9EAD3")
thin = Side(style="thin", color="D9E2F3")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for ws in wb.worksheets:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = True

    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

            if ws.title == "quality_gates":
                if cell.column == 2:
                    if str(cell.value).lower() == "pass":
                        cell.fill = pass_fill
                    elif str(cell.value).lower() == "warning":
                        cell.fill = warning_fill
                    elif str(cell.value).lower() == "fail":
                        cell.fill = fail_fill

            if ws.title == "claims":
                if cell.column == 1:
                    if str(cell.value).lower() == "allowed":
                        cell.fill = pass_fill
                    elif str(cell.value).lower() == "avoid":
                        cell.fill = fail_fill

    # Auto-width with sensible caps
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = 0
        col_letter = get_column_letter(col_idx)

        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        width = min(max(max_length + 2, 10), 48)
        ws.column_dimensions[col_letter].width = width

    # Number formats
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.000"

# Make README easier to read
if "README" in wb.sheetnames:
    ws = wb["README"]
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 90

wb.save(OUTPUT_PATH)


# ------------------------------------------------------------
# 14) Console summary
# ------------------------------------------------------------

print("\nCreated:", OUTPUT_PATH.resolve())
print("\nSheets written:")
for name in sheets:
    print(f"- {name}: {len(sheets[name])} rows")

print("\nQuality gates:")
print(quality_gates.to_string(index=False))

print("\nNext recommended files to review/upload:")
print("- phase12_results_master_tables.xlsx")
print("\nInterpretation reminder:")
print("This workbook is a manuscript-ready evidence package, not a new modeling phase.")
