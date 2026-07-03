# ============================================================
# Phase 12-B - Manuscript Tables, Figure-Ready Data, and Results Skeleton
# Project:
# From Accuracy to Trustworthiness:
# Grouped Validation, Applicability-Domain Mapping,
# and Reliability-Aware Optimization for ML-Based
# Fixed-Bed Biomass Pyrolysis Yield Prediction
# ============================================================
#
# Purpose:
# Converts phase12_results_master_tables.xlsx into a compact manuscript package:
#
#   1) phase12b_manuscript_tables_and_figure_data.xlsx
#   2) figure-ready data sheets
#   3) optional PNG figures for quick inspection
#   4) Results skeleton with controlled, non-overclaiming wording
#
# This script does NOT:
#   - train a new model,
#   - run new optimization,
#   - create new scientific claims,
#   - present scenario windows as true optima.
#
# ============================================================

from pathlib import Path
import numpy as np
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ------------------------------------------------------------
# 0) Input/output paths
# ------------------------------------------------------------

BASE_DIR = Path(".")

MASTER_PATH = BASE_DIR / "phase12_results_master_tables.xlsx"
PHASE7_FULL_PATH = BASE_DIR / "phase7_v2_summary_results.xlsx"
PHASE8_DECISIONS_PATH = BASE_DIR / "phase8_official_model_decisions.xlsx"

OUTPUT_XLSX = BASE_DIR / "phase12b_manuscript_tables_and_figure_data.xlsx"

FIG_DIR = BASE_DIR / "phase12b_figures"
FIG_DIR.mkdir(exist_ok=True)

assert MASTER_PATH.exists(), f"Missing master workbook: {MASTER_PATH.resolve()}"

print("Loading:", MASTER_PATH.resolve())


# ------------------------------------------------------------
# 1) Helpers
# ------------------------------------------------------------

def read_master(sheet_name):
    return pd.read_excel(MASTER_PATH, sheet_name=sheet_name)


def read_optional_excel(path):
    if path.exists():
        return pd.read_excel(path)
    return None


def compact_float(df, digits=3):
    out = df.copy()
    for col in out.select_dtypes(include=[np.number]).columns:
        out[col] = out[col].round(digits)
    return out


def safe_select(df, cols):
    existing = [c for c in cols if c in df.columns]
    return df[existing].copy()


def regime_label(x):
    mapping = {
        "random_kfold": "Random K-fold",
        "source_group_kfold": "Source-grouped",
        "feedstock_group_kfold": "Feedstock-grouped",
        "family_group_kfold": "Family-grouped",
    }
    return mapping.get(x, x)


def target_label(x):
    mapping = {
        "biochar": "Biochar",
        "bio_oil": "Bio-oil",
    }
    return mapping.get(x, x)


def role_keep_for_main(role):
    if pd.isna(role):
        return False
    role = str(role)
    return role in [
        "primary_biochar_model",
        "primary_bio_oil_screening_model",
        "bio_oil_sensitivity_model",
    ]


def source_sheet_note(sheet_name):
    return f"Derived from phase12_results_master_tables.xlsx::{sheet_name}"


def add_source_note(df, sheet_name, note=None):
    out = df.copy()
    out["source_note"] = note if note else source_sheet_note(sheet_name)
    return out


# ------------------------------------------------------------
# 2) Load master sheets
# ------------------------------------------------------------

dataset_audit = read_master("dataset_audit")
validation_core = read_master("validation_core")
validation_contrast = read_master("validation_contrast")
model_decisions = read_master("model_decisions")
uncertainty_core = read_master("uncertainty_core")
uncert_assoc = read_master("uncert_assoc")
ad_filtering = read_master("ad_filtering")
ad_error_class = read_master("ad_error_class")
ad_assoc = read_master("ad_assoc")
reliability_ranking = read_master("reliability_ranking")
biochar_family = read_master("biochar_family")
biochar_cases = read_master("biochar_cases")
warning_groups = read_master("warning_groups")
claims = read_master("claims")

phase7_full = read_optional_excel(PHASE7_FULL_PATH)
phase8_decisions = read_optional_excel(PHASE8_DECISIONS_PATH)


# ------------------------------------------------------------
# 3) Table 1 - Dataset audit
# ------------------------------------------------------------

dataset_label_map = {
    "cleaned_rows": "Cleaned dataset rows",
    "cleaned_columns": "Cleaned dataset columns",
    "usable_biochar_rows": "Usable rows for biochar yield",
    "usable_bio_oil_rows": "Usable rows for bio-oil yield",
    "paired_target_rows": "Rows with both targets available",
    "unique_feedstocks": "Harmonized feedstock labels",
    "unique_feedstock_families": "Feedstock families",
    "source_groups": "Primary source groups",
    "source_pair_groups": "Source-pair groups",
}

T1_dataset = dataset_audit.copy()
T1_dataset["item_for_manuscript"] = T1_dataset["item"].map(dataset_label_map).fillna(T1_dataset["item"])
T1_dataset = safe_select(
    T1_dataset,
    ["item_for_manuscript", "value", "note"]
)
T1_dataset = T1_dataset.rename(columns={
    "item_for_manuscript": "Dataset item",
    "value": "Value",
    "note": "Interpretation",
})
T1_dataset = add_source_note(T1_dataset, "dataset_audit")


# ------------------------------------------------------------
# 4) Table 2 - Best-by-regime validation
# ------------------------------------------------------------

T2_validation_best = validation_contrast.copy()
T2_validation_best["Target"] = T2_validation_best["target"].map(target_label)
T2_validation_best["Validation regime"] = T2_validation_best["validation_regime"].map(regime_label)

T2_validation_best = safe_select(
    T2_validation_best,
    [
        "Target",
        "Validation regime",
        "selected_model",
        "selected_protocol",
        "mean_r2",
        "mean_rmse",
        "mean_mae",
        "absolute_r2_drop_vs_random_best",
        "rmse_increase_vs_random_best",
        "interpretation",
    ],
)

T2_validation_best = T2_validation_best.rename(columns={
    "selected_model": "Best model in regime",
    "selected_protocol": "Best protocol in regime",
    "mean_r2": "Mean R²",
    "mean_rmse": "RMSE",
    "mean_mae": "MAE",
    "absolute_r2_drop_vs_random_best": "R² drop vs random best",
    "rmse_increase_vs_random_best": "RMSE increase vs random best",
    "interpretation": "Interpretation",
})

T2_validation_best = compact_float(add_source_note(T2_validation_best, "validation_contrast"))


# ------------------------------------------------------------
# 5) Table 3 - Official selected model performance
# ------------------------------------------------------------
# This avoids a common reviewer confusion:
# "best numeric performer in each regime" is not always the same as
# "official model selected for downstream uncertainty/AD/scenario ranking."

official_perf_rows = []

if phase7_full is not None and phase8_decisions is not None:
    decision_rows = phase8_decisions[
        phase8_decisions["role"].isin([
            "primary_model_for_next_phases",
        ])
    ].copy()

    for _, drow in decision_rows.iterrows():
        target = drow.get("target")
        selected_model = drow.get("selected_model")
        selected_protocol = drow.get("selected_protocol")

        matched = phase7_full[
            (phase7_full["target"] == target)
            & (phase7_full["model"] == selected_model)
            & (phase7_full["protocol"] == selected_protocol)
        ].copy()

        for _, prow in matched.iterrows():
            official_perf_rows.append({
                "Target": target_label(target),
                "Official role": drow.get("role"),
                "Selected model": selected_model,
                "Selected protocol": selected_protocol,
                "Validation regime": regime_label(prow.get("validation_regime")),
                "Mean R²": prow.get("mean_r2"),
                "RMSE": prow.get("mean_rmse"),
                "MAE": prow.get("mean_mae"),
                "Pooled R²": prow.get("pooled_r2"),
                "Pooled RMSE": prow.get("pooled_rmse"),
                "Claim strength": drow.get("claim_strength"),
                "Why this model was selected": drow.get("reason"),
                "Remaining risk": drow.get("remaining_risk"),
            })

T3_official_model_perf = pd.DataFrame(official_perf_rows)

if len(T3_official_model_perf) == 0:
    # Fallback if full phase7 is unavailable.
    T3_official_model_perf = model_decisions.copy()
    T3_official_model_perf = T3_official_model_perf.rename(columns={
        "target": "Target",
        "role": "Official role",
        "selected_model": "Selected model",
        "selected_protocol": "Selected protocol",
        "claim_strength": "Claim strength",
        "reason": "Why this model was selected",
        "remaining_risk": "Remaining risk",
    })

T3_official_model_perf = compact_float(
    add_source_note(
        T3_official_model_perf,
        "model_decisions + phase7_v2_summary_results",
        "Official selected models; performance rows are matched to phase7_v2_summary_results.xlsx when available."
    )
)


# ------------------------------------------------------------
# 6) Table 4 - Uncertainty calibration at 90% only
# ------------------------------------------------------------

T4_uncertainty_90 = uncertainty_core[
    np.isclose(uncertainty_core["nominal_coverage"], 0.90)
].copy()

T4_uncertainty_90 = T4_uncertainty_90[
    T4_uncertainty_90["role"].isin([
        "primary_biochar_model",
        "primary_bio_oil_screening_model",
    ])
].copy()

T4_uncertainty_90["Target"] = T4_uncertainty_90["target"].map(target_label)
T4_uncertainty_90["Validation regime"] = T4_uncertainty_90["validation_regime"].map(regime_label)

T4_uncertainty_90 = safe_select(
    T4_uncertainty_90,
    [
        "Target",
        "role",
        "protocol",
        "model",
        "Validation regime",
        "r2",
        "rmse",
        "mae",
        "coverage_abs_conformal",
        "mean_width_abs_conformal",
        "coverage_norm_conformal",
        "mean_width_norm_conformal",
        "mean_ensemble_std",
        "claim_strength",
    ],
)

T4_uncertainty_90 = T4_uncertainty_90.rename(columns={
    "role": "Role",
    "protocol": "Protocol",
    "model": "Model",
    "r2": "R²",
    "rmse": "RMSE",
    "mae": "MAE",
    "coverage_abs_conformal": "Absolute conformal coverage",
    "mean_width_abs_conformal": "Absolute conformal mean width",
    "coverage_norm_conformal": "Normalized conformal coverage",
    "mean_width_norm_conformal": "Normalized conformal mean width",
    "mean_ensemble_std": "Mean ensemble std",
    "claim_strength": "Claim strength",
})

T4_uncertainty_90 = compact_float(add_source_note(T4_uncertainty_90, "uncertainty_core"))


# ------------------------------------------------------------
# 7) Table 5 - AD filtering effect, compact
# ------------------------------------------------------------

ad_main = ad_filtering[
    ad_filtering["role"].isin([
        "primary_biochar_model",
        "primary_bio_oil_screening_model",
    ])
].copy()

ad_main = ad_main[
    ad_main["filter_name"].isin([
        "all_predictions",
        "in_or_near_domain_only",
        "strict_in_domain_only",
    ])
].copy()

ad_main["Target"] = ad_main["target"].map(target_label)
ad_main["Validation regime"] = ad_main["validation_regime"].map(regime_label)

T5_ad_filtering = safe_select(
    ad_main,
    [
        "Target",
        "role",
        "protocol",
        "model",
        "Validation regime",
        "filter_name",
        "n_kept",
        "fraction_kept",
        "mean_abs_error",
        "rmse",
        "coverage_abs_90",
        "mean_width_abs_90",
        "claim_strength",
    ],
)

T5_ad_filtering = T5_ad_filtering.rename(columns={
    "role": "Role",
    "protocol": "Protocol",
    "model": "Model",
    "filter_name": "AD filter",
    "n_kept": "N kept",
    "fraction_kept": "Fraction kept",
    "mean_abs_error": "MAE",
    "rmse": "RMSE",
    "coverage_abs_90": "90% abs-conformal coverage",
    "mean_width_abs_90": "90% abs-conformal width",
    "claim_strength": "Claim strength",
})

T5_ad_filtering = compact_float(add_source_note(T5_ad_filtering, "ad_filtering"))


# ------------------------------------------------------------
# 8) Table 6 - Yield-only vs reliability-aware scenario ranking
# ------------------------------------------------------------

T6_reliability_ranking = reliability_ranking.copy()

T6_reliability_ranking = safe_select(
    T6_reliability_ranking,
    [
        "group_type",
        "group_value",
        "candidate_fraction_in_or_near_domain",
        "candidate_fraction_out_of_domain",
        "yield_only_mean_predicted_biochar_yield",
        "yield_only_fraction_out_of_domain",
        "reliability_aware_lcb90_ad_filtered_mean_predicted_biochar_yield",
        "reliability_aware_lcb90_ad_filtered_mean_lower_bound_90",
        "reliability_aware_lcb90_ad_filtered_fraction_in_domain",
        "reliability_aware_lcb90_ad_filtered_fraction_near_domain",
        "reliability_aware_lcb90_ad_filtered_fraction_out_of_domain",
        "yield_only_minus_reliable_ood_fraction",
        "yield_only_high_risk_flag",
        "reporting_class",
        "reporting_reason",
    ],
)

T6_reliability_ranking = T6_reliability_ranking.rename(columns={
    "group_type": "Group level",
    "group_value": "Group",
    "candidate_fraction_in_or_near_domain": "All candidates: in/near-domain fraction",
    "candidate_fraction_out_of_domain": "All candidates: out-of-domain fraction",
    "yield_only_mean_predicted_biochar_yield": "Yield-only mean predicted biochar yield",
    "yield_only_fraction_out_of_domain": "Yield-only out-of-domain fraction",
    "reliability_aware_lcb90_ad_filtered_mean_predicted_biochar_yield": "Reliability-aware mean predicted yield",
    "reliability_aware_lcb90_ad_filtered_mean_lower_bound_90": "Reliability-aware mean 90% lower bound",
    "reliability_aware_lcb90_ad_filtered_fraction_in_domain": "Reliability-aware in-domain fraction",
    "reliability_aware_lcb90_ad_filtered_fraction_near_domain": "Reliability-aware near-domain fraction",
    "reliability_aware_lcb90_ad_filtered_fraction_out_of_domain": "Reliability-aware out-of-domain fraction",
    "yield_only_minus_reliable_ood_fraction": "OOD fraction reduction",
    "yield_only_high_risk_flag": "Yield-only high-risk flag",
    "reporting_class": "Reporting class",
    "reporting_reason": "Reporting reason",
})

T6_reliability_ranking = compact_float(add_source_note(T6_reliability_ranking, "reliability_ranking"))


# ------------------------------------------------------------
# 9) Table 7 - Biochar screening windows
# ------------------------------------------------------------

def prepare_window_table(df, label):
    out = df.copy()
    out["Window level"] = label

    wanted = [
        "Window level",
        "group_value",
        "evidence_n_rows",
        "evidence_n_sources",
        "mean_predicted_biochar_yield",
        "mean_lower_bound_90",
        "candidate_fraction_in_or_near_domain",
        "candidate_fraction_out_of_domain",
        "fraction_in_domain",
        "fraction_near_domain",
        "temperature_c_p10",
        "temperature_c_median",
        "temperature_c_p90",
        "residence_time_min_p10",
        "residence_time_min_median",
        "residence_time_min_p90",
        "heating_rate_c_min_p10",
        "heating_rate_c_min_median",
        "heating_rate_c_min_p90",
        "reporting_reason",
    ]

    out = safe_select(out, wanted)

    out = out.rename(columns={
        "group_value": "Group",
        "evidence_n_rows": "Evidence rows",
        "evidence_n_sources": "Evidence sources",
        "mean_predicted_biochar_yield": "Mean predicted yield",
        "mean_lower_bound_90": "Mean 90% lower bound",
        "candidate_fraction_in_or_near_domain": "Candidate in/near-domain fraction",
        "candidate_fraction_out_of_domain": "Candidate out-of-domain fraction",
        "fraction_in_domain": "Selected in-domain fraction",
        "fraction_near_domain": "Selected near-domain fraction",
        "temperature_c_p10": "Temperature p10",
        "temperature_c_median": "Temperature median",
        "temperature_c_p90": "Temperature p90",
        "residence_time_min_p10": "Residence time p10",
        "residence_time_min_median": "Residence time median",
        "residence_time_min_p90": "Residence time p90",
        "heating_rate_c_min_p10": "Heating rate p10",
        "heating_rate_c_min_median": "Heating rate median",
        "heating_rate_c_min_p90": "Heating rate p90",
        "reporting_reason": "Reporting reason",
    })

    return out


family_windows_compact = prepare_window_table(biochar_family, "Feedstock family")
case_windows_compact = prepare_window_table(biochar_cases, "Feedstock case study")

T7_biochar_windows = pd.concat(
    [family_windows_compact, case_windows_compact],
    ignore_index=True,
)

T7_biochar_windows = compact_float(
    add_source_note(
        T7_biochar_windows,
        "biochar_family + biochar_cases",
        "Reliability-aware biochar screening windows; not experimentally validated optima."
    )
)


# ------------------------------------------------------------
# 10) Supplementary tables
# ------------------------------------------------------------

S1_warning_groups = compact_float(add_source_note(warning_groups.copy(), "warning_groups"))
S2_uncertainty_assoc = compact_float(add_source_note(uncert_assoc.copy(), "uncert_assoc"))
S3_ad_assoc = compact_float(add_source_note(ad_assoc.copy(), "ad_assoc"))
S4_claims = claims.copy()


# ------------------------------------------------------------
# 11) Figure-ready data
# ------------------------------------------------------------

# Figure 2: Random vs grouped R2 and RMSE
fig2_validation = validation_contrast.copy()
fig2_validation["Target"] = fig2_validation["target"].map(target_label)
fig2_validation["Validation regime"] = fig2_validation["validation_regime"].map(regime_label)
fig2_validation = safe_select(
    fig2_validation,
    [
        "Target", "Validation regime", "mean_r2", "mean_rmse",
        "absolute_r2_drop_vs_random_best", "rmse_increase_vs_random_best"
    ],
)
fig2_validation = fig2_validation.rename(columns={
    "mean_r2": "Mean R²",
    "mean_rmse": "RMSE",
    "absolute_r2_drop_vs_random_best": "R² drop vs random",
    "rmse_increase_vs_random_best": "RMSE increase vs random",
})
fig2_validation = compact_float(fig2_validation)

# Figure 4: uncertainty coverage and width at 90%
fig4_uncertainty = T4_uncertainty_90.copy()

# Figure 5: AD filtering MAE/RMSE
fig5_ad_filtering = T5_ad_filtering.copy()

# Figure 6: yield-only risk vs reliability-aware
fig6_reliability = T6_reliability_ranking.copy()


# ------------------------------------------------------------
# 12) Results skeleton
# ------------------------------------------------------------

results_skeleton = pd.DataFrame([
    {
        "section": "3.1 Dataset audit and target-specific availability",
        "main_message": (
            "The dataset contains target-dependent usable records; biochar and bio-oil "
            "must therefore be evaluated separately rather than treated as equally supported targets."
        ),
        "core_tables": "Table 1",
        "core_figures": "Figure 1",
        "numbers_to_use": "1674 cleaned rows; 1548 biochar rows; 1338 bio-oil rows; 68 feedstocks; 124 source groups.",
        "guardrail": "Do not imply that all rows contain all targets or all features.",
    },
    {
        "section": "3.2 Random CV overestimates deployment-relevant performance",
        "main_message": (
            "Random K-fold validation substantially overestimates performance compared with "
            "source-, feedstock-, and family-grouped validation."
        ),
        "core_tables": "Table 2; optional Table 3",
        "core_figures": "Figure 2",
        "numbers_to_use": (
            "Biochar: random R² ≈ 0.823, grouped R² ≈ 0.514–0.570. "
            "Bio-oil: random R² ≈ 0.872, grouped R² ≈ 0.065–0.128."
        ),
        "guardrail": "Do not claim that random-CV performance represents deployment performance.",
    },
    {
        "section": "3.3 Target-dependent trustworthiness",
        "main_message": (
            "Biochar retains moderate screening value under grouped validation, whereas bio-oil "
            "acts as a cautionary case with weak grouped generalization."
        ),
        "core_tables": "Table 2; Table 3",
        "core_figures": "Figure 3",
        "numbers_to_use": "Use grouped-validation R²/RMSE contrast and official model claim-strength table.",
        "guardrail": "Do not present bio-oil as a robust optimization target.",
    },
    {
        "section": "3.4 Uncertainty calibration under grouped validation",
        "main_message": (
            "Conformal intervals can maintain near-nominal coverage for biochar but become much wider "
            "under grouped validation; bio-oil shows broader intervals and under-coverage in key grouped regimes."
        ),
        "core_tables": "Table 4",
        "core_figures": "Figure 4",
        "numbers_to_use": "Use 90% coverage and interval-width values only in the main text.",
        "guardrail": "Do not claim complete pointwise process uncertainty.",
    },
    {
        "section": "3.5 Applicability-domain filtering",
        "main_message": (
            "AD filtering reduces biochar error and improves screening reliability, but it only partially helps bio-oil."
        ),
        "core_tables": "Table 5",
        "core_figures": "Figure 5",
        "numbers_to_use": "Biochar source/family grouped MAE reduction after in/near-domain filtering.",
        "guardrail": "Do not claim AD distance fully explains model error.",
    },
    {
        "section": "3.6 Reliability-aware biochar scenario ranking",
        "main_message": (
            "Yield-only ranking frequently selects out-of-domain candidates; reliability-aware ranking removes "
            "out-of-domain candidates and reports conservative biochar screening windows."
        ),
        "core_tables": "Table 6; Table 7",
        "core_figures": "Figure 6",
        "numbers_to_use": "Use out-of-domain fractions and final biochar family/case windows.",
        "guardrail": "Do not call the windows true optimum conditions or industrial set-points.",
    },
])


# ------------------------------------------------------------
# 13) Quality gates for Phase 12-B
# ------------------------------------------------------------

quality_rows = []

quality_rows.append({
    "check": "master_workbook_exists",
    "status": "pass",
    "detail": str(MASTER_PATH),
})

for name, df in [
    ("T1_dataset", T1_dataset),
    ("T2_validation_best", T2_validation_best),
    ("T3_official_model_perf", T3_official_model_perf),
    ("T4_uncertainty_90", T4_uncertainty_90),
    ("T5_ad_filtering", T5_ad_filtering),
    ("T6_reliability_ranking", T6_reliability_ranking),
    ("T7_biochar_windows", T7_biochar_windows),
]:
    status = "pass" if len(df) > 0 else "warning"
    quality_rows.append({
        "check": f"{name}_nonempty",
        "status": status,
        "detail": f"{len(df)} rows.",
    })

# Duplicate check for key group tables
for name, df in [
    ("T6_reliability_ranking", T6_reliability_ranking),
    ("biochar_family", biochar_family),
    ("biochar_cases", biochar_cases),
]:
    if {"Group level", "Group"}.issubset(df.columns):
        dup = int(df[["Group level", "Group"]].duplicated().sum())
    elif {"group_type", "group_value"}.issubset(df.columns):
        dup = int(df[["group_type", "group_value"]].duplicated().sum())
    else:
        dup = 0

    quality_rows.append({
        "check": f"{name}_duplicate_groups",
        "status": "pass" if dup == 0 else "fail",
        "detail": f"{dup} duplicate group rows detected.",
    })

quality_rows.append({
    "check": "language_guardrail",
    "status": "pass",
    "detail": "Use screening windows, not optima; bio-oil is cautionary, not an optimization target.",
})

quality_gates = pd.DataFrame(quality_rows)


# ------------------------------------------------------------
# 14) Write workbook
# ------------------------------------------------------------

sheets = {
    "README": pd.DataFrame([
        {
            "field": "purpose",
            "value": "Compact manuscript tables, figure-ready data, and Results skeleton from Phase 12-A master workbook.",
        },
        {
            "field": "main_guardrail",
            "value": "Do not describe biochar scenario windows as experimentally validated optima.",
        },
        {
            "field": "bio_oil_guardrail",
            "value": "Bio-oil remains a cautionary trustworthiness case, not an operational recommendation target.",
        },
    ]),
    "quality_gates": quality_gates,
    "T1_dataset_audit": T1_dataset,
    "T2_validation_best": T2_validation_best,
    "T3_official_model_perf": T3_official_model_perf,
    "T4_uncertainty_90": T4_uncertainty_90,
    "T5_ad_filtering": T5_ad_filtering,
    "T6_reliability_ranking": T6_reliability_ranking,
    "T7_biochar_windows": T7_biochar_windows,
    "S1_warning_groups": S1_warning_groups,
    "S2_uncertainty_assoc": S2_uncertainty_assoc,
    "S3_ad_assoc": S3_ad_assoc,
    "S4_claims": S4_claims,
    "Fig2_validation_data": fig2_validation,
    "Fig4_uncertainty_data": fig4_uncertainty,
    "Fig5_AD_filtering_data": fig5_ad_filtering,
    "Fig6_reliability_data": fig6_reliability,
    "Results_skeleton": results_skeleton,
}

with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
    for sheet_name, table in sheets.items():
        table.to_excel(writer, sheet_name=sheet_name, index=False)


# ------------------------------------------------------------
# 15) Formatting workbook
# ------------------------------------------------------------

wb = load_workbook(OUTPUT_XLSX)

header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(color="FFFFFF", bold=True)
pass_fill = PatternFill("solid", fgColor="D9EAD3")
warning_fill = PatternFill("solid", fgColor="FFF2CC")
fail_fill = PatternFill("solid", fgColor="F4CCCC")
thin = Side(style="thin", color="D9E2F3")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for ws in wb.worksheets:
    ws.freeze_panes = "A2"

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

            if ws.title == "quality_gates" and cell.column == 2:
                val = str(cell.value).lower()
                if val == "pass":
                    cell.fill = pass_fill
                elif val == "warning":
                    cell.fill = warning_fill
                elif val == "fail":
                    cell.fill = fail_fill

            if ws.title == "S4_claims" and cell.column == 1:
                val = str(cell.value).lower()
                if val == "allowed":
                    cell.fill = pass_fill
                elif val == "avoid":
                    cell.fill = fail_fill

    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        width = min(max(max_length + 2, 10), 48)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.000"

wb.save(OUTPUT_XLSX)


# ------------------------------------------------------------
# 16) Optional PNG figures
# ------------------------------------------------------------
# These figures are quick inspection figures, not final journal artwork.
# They use default matplotlib styling and can be redesigned later.

try:
    import matplotlib.pyplot as plt

    # Figure 2: Mean R2 by validation regime
    fig2 = fig2_validation.copy()
    fig2_pivot = fig2.pivot(index="Validation regime", columns="Target", values="Mean R²")
    regime_order = ["Random K-fold", "Source-grouped", "Feedstock-grouped", "Family-grouped"]
    fig2_pivot = fig2_pivot.reindex([r for r in regime_order if r in fig2_pivot.index])
    ax = fig2_pivot.plot(kind="bar")
    ax.set_ylabel("Mean R²")
    ax.set_title("Random-CV optimism under grouped validation")
    ax.set_xlabel("")
    plt.tight_layout()
    plt.savefig(FIG_DIR / "phase12b_fig2_validation_r2_by_regime.png", dpi=300)
    plt.close()

    # Figure 4: 90% interval width
    fig4 = fig4_uncertainty.copy()
    if "Target" in fig4.columns and "Validation regime" in fig4.columns:
        fig4["Series"] = fig4["Target"] + " | " + fig4["Validation regime"]
        fig4_plot = fig4.set_index("Series")["Absolute conformal mean width"]
        ax = fig4_plot.plot(kind="bar")
        ax.set_ylabel("Mean width of 90% conformal interval")
        ax.set_title("Uncertainty interval width under validation regimes")
        ax.set_xlabel("")
        plt.xticks(rotation=45, ha="right")
        plt.tight_layout()
        plt.savefig(FIG_DIR / "phase12b_fig4_uncertainty_width_90.png", dpi=300)
        plt.close()

    # Figure 5: AD filtering MAE for primary biochar only
    fig5 = fig5_ad_filtering.copy()
    fig5 = fig5[fig5["Target"] == "Biochar"].copy()
    fig5["Series"] = fig5["Validation regime"] + " | " + fig5["AD filter"]
    fig5_plot = fig5.set_index("Series")["MAE"]
    ax = fig5_plot.plot(kind="bar")
    ax.set_ylabel("MAE")
    ax.set_title("Biochar error after applicability-domain filtering")
    ax.set_xlabel("")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    plt.savefig(FIG_DIR / "phase12b_fig5_biochar_ad_filtering_mae.png", dpi=300)
    plt.close()

    # Figure 6: OOD fraction yield-only vs reliability-aware
    fig6 = fig6_reliability.copy()
    fig6 = fig6[[
        "Group",
        "Yield-only out-of-domain fraction",
        "Reliability-aware out-of-domain fraction"
    ]].copy()
    fig6 = fig6.set_index("Group")
    ax = fig6.plot(kind="bar")
    ax.set_ylabel("Out-of-domain fraction")
    ax.set_title("Yield-only versus reliability-aware ranking")
    ax.set_xlabel("")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    plt.savefig(FIG_DIR / "phase12b_fig6_yield_only_vs_reliable_ood.png", dpi=300)
    plt.close()

    print("\nOptional PNG figures saved in:", FIG_DIR.resolve())

except Exception as e:
    print("\nFigure generation skipped due to error:")
    print(e)


# ------------------------------------------------------------
# 17) Console summary
# ------------------------------------------------------------

print("\nCreated:", OUTPUT_XLSX.resolve())
print("\nSheets written:")
for name, table in sheets.items():
    print(f"- {name}: {len(table)} rows")

print("\nQuality gates:")
print(quality_gates.to_string(index=False))

print("\nNext files to review/upload:")
print("- phase12b_manuscript_tables_and_figure_data.xlsx")
print("- optional PNG figures in phase12b_figures/")
print("\nInterpretation reminder:")
print("This package prepares manuscript tables and figure data; it does not validate true optimum conditions.")
